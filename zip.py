<<<<<<< HEAD
=======
#TASKS:-


#Make it neat
##Make script efficient and eliminate complexity
#add comments explaining what has been done
#Apply Datastructures and algorithms where required
#Use OOPs concepts
#Make it pythonic
#use modules if required

>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
import os
import ipaddress
import openpyxl
import valid_ip  # imported module valid_ip.py
from openpyxl.styles import PatternFill, Font
# from openpyxl.styles import *
from openpyxl.utils.cell import get_column_letter
<<<<<<< HEAD

ab = 1.2
c_green_a = [0, 1, 2, 3, 4, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 91, 92, 93, 94, 96, 97,
             98, 99, 100, 101, 102, 103, 104, 108]
c_grey_a = [42, 43, 120, 121]
c_yellow_a = [11, 12, 89, 90]
c_dict_a = {'green': c_green_a, 'grey': c_grey_a, 'yellow': c_yellow_a}
d_green_41_95_start_a = 128
d_green_41_95_end_a = 255
d_grey_44_start_a = 0
d_grey_44_end_a = 159
d_grey_122_start_a = 0
d_grey_122_end_a = 95

# Sheet2 subnets
c_green_b = [141, 142, 143, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 156, 157, 169]
c_grey_b = [170, 171, 172]
c_yellow_b = [139, 140]
c_dict_b = {'green': c_green_b, 'grey': c_grey_b, 'yellow': c_yellow_b}
d_green_169_start_b = 0
d_green_169_end_b = 127
d_grey_172_1_start_b = 0
d_grey_172_1_end_b = 63
d_grey_172_2_start_b = 128
d_grey_172_2_end_b = 159

bluezone = ['bluezone', 'Bluezone']

common_apps = {
    'ssl': '443', 'web-browsing': '80', 'dns': '53', 'ftp': '21', 'ssh': '22', 'netbios-ss': '139',
    'snmp': '161', 'ms-rdp': '3389', 'socks': '1080', 'ntp': '123', 'tacacs': '49', 'telnet': '23', 'portmapper': '111',
    'nfs': '2049', 'mysql': '3306', 'wins': '42', 'rdp2tcp': '3389', 'lotus-notes-base': '1352', 'smtp': '25'
}

common_apps_list = {
    'ldap': ['389', '3268', '3269', '636'], 'db2': ['50000', '60000'], 'ms-ds-smb': ['139', '445'],
    'kerberos': ['88', '464', '749', '750', '754'], 'ms-netlogon': ['1025-5000', '49152-65535', '137', '138'],
    'WinRM': ['5985', '5986'],
    'windows-remote-management': ['5985', '5986']
}

=======
ab=1.2
c_green_a=[0,1,2,3,4,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,91,92,93,94,96,97,98,99,100,101,102,103,104,108]
c_grey_a = [42, 43, 120, 121]
c_yellow_a = [11, 12, 89, 90]
c_dict_a = {'green': c_green_a, 'grey': c_grey_a, 'yellow': c_yellow_a}
d_green_41_95_start_a=128
d_green_41_95_end_a=255
d_grey_44_start_a=0
d_grey_44_end_a=159
d_grey_122_start_a=0
d_grey_122_end_a=95

#Sheet2 subnets
c_green_b=[141,142,143,144,145,146,147,148,149,150,151,152,153,154,155,156,157,169]
c_grey_b= [170,171,172]
c_yellow_b= [139,140]
c_dict_b = {'green': c_green_b, 'grey': c_grey_b, 'yellow': c_yellow_b}
d_green_169_start_b=0
d_green_169_end_b=127
d_grey_172_1_start_b=0
d_grey_172_1_end_b=63
d_grey_172_2_start_b = 128
d_grey_172_2_end_b = 159

bluezone=['bluezone','Bluezone']

common_apps = {
        'ssl': '443', 'web-browsing': '80', 'dns': '53', 'ftp': '21', 'ssh': '22', 'netbios-ss': '139',
        'snmp': '161', 'ms-rdp': '3389', 'socks': '1080', 'ntp': '123', 'tacacs': '49', 'telnet': '23', 'portmapper': '111', 'nfs': '2049', 'mysql': '3306',  'wins': '42', 'rdp2tcp': '3389', 'lotus-notes-base': '1352', 'smtp': '25'
    }

common_apps_list = {
        'ldap': ['389', '3268', '3269', '636'], 'db2': ['50000', '60000'], 'ms-ds-smb': ['139', '445'], 'kerberos': ['88', '464', '749', '750', '754'], 'ms-netlogon': ['1025-5000', '49152-65535', '137', '138'], 'WinRM': ['5985', '5986'],
        'windows-remote-management': ['5985', '5986']
    }
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8

def get_key(val, search_dict):
    """
    Gets dict key from supplied value.
    val: Value to search
    search_dict : Dictionary to search value for
<<<<<<< HEAD
=======

>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
    """
    for key, value in search_dict.items():
        if val in value:
            return key

<<<<<<< HEAD

def check_zone_ip_relationship(inp_ip, inp):
=======
def is_valid(ip):
    """
    Checks if IP address is valid or not.
    Input: IP address
    Output: bool 
    """
    try:
        ipaddress.IPv4Network(ip)
        return True
    except ValueError:
        return False


def check_zone_ip_relationship(inp_ip,inp): 
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
    """
    checking which zone does the ip belongs to based on the excel sheet at Hacktoberfest_database.xlsx
    Input: 
    inp_ip : Input IP address
    inp: (int) Sheet number
    Output:
    color value
    """
<<<<<<< HEAD
    ss = os.path.join(os.path.dirname(__file__), 'Hacktoberfest_database.xlsx')
    load_spreadsheet = openpyxl.load_workbook(ss)
    a_sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
    b_sheet = load_spreadsheet['Sheet2']  # Get a sheet from the workbook

    if inp.lower() == 'sheet1' or inp == '1':
        # Operations for sheet 1
        for row in a_sheet.iter_cols(min_col=1, min_row=4, max_col=4,
                                     max_row=100):  # iterate through all rows in specific column
            for cell in row:
                if cell.value != None:
                    if '/' not in inp_ip:
                        abcd = inp_ip.split('.')
                    else:
                        remove_network = inp_ip.split('/')
                        abcd = remove_network[0].split('.')
=======
    ss=os.path.join(os.path.dirname(__file__),'Hacktoberfest_database.xlsx')
    load_spreadsheet = openpyxl.load_workbook(ss)
    a_sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
    b_sheet = load_spreadsheet['Sheet2']  # Get a sheet from the workbook
    
    if inp.lower()=='sheet1' or inp=='1': 
        # Operations for sheet 1
        for row in a_sheet.iter_cols(min_col=1, min_row=4, max_col=4,max_row=100):  # iterate through all rows in specific column
            for cell in row:
                if cell.value!=None:
                    if '/' not in inp_ip:
                        abcd=inp_ip.split('.')
                    else:
                        remove_network=inp_ip.split('/')
                        abcd=remove_network[0].split('.')
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                    if abcd[0] != ('1'):
                        return ('red')
                    elif inp_ip in bluezone or float(abcd[0] + '.' + abcd[1]) != ab:
                        return ('blue')
<<<<<<< HEAD
                    elif float(abcd[0] + '.' + abcd[1]) == ab:

                        c_key = get_key(abcd[2], c_dict_a)

                        if int(abcd[2]) == int('41') or int(abcd[2]) == int('95'):
                            if int(abcd[3]) in range(d_green_41_95_start_a, d_green_41_95_end_a + 1):
                                return ('green')
                        elif int(abcd[2]) == int(44):
                            for i in range(d_grey_44_start_a, d_grey_44_end_a + 1):
                                if i == int(abcd[3]):
                                    return ('grey')
                        elif int(abcd[2]) == int(122):
                            for i in range(d_grey_122_start_a, d_grey_122_end_a + 1):
                                if i == int(abcd[3]):
                                    return ('grey')
                        return c_key

    if inp.lower() == 'sheet2' or inp == '2':
        # Operations for sheet 2
        for row in b_sheet.iter_cols(min_col=1, min_row=4, max_col=4,
                                     max_row=100):  # iterate through all rows in specific column
=======
                    elif float(abcd[0]+'.'+abcd[1])==ab:

                        c_key = get_key(abcd[2], c_dict_a)
                        
                        if int(abcd[2])==int('41') or int(abcd[2])==int('95'):
                            if int(abcd[3]) in range(d_green_41_95_start_a,d_green_41_95_end_a+1):

                                return ('green')
                        elif int(abcd[2])==int(44):
                            for i in range(d_grey_44_start_a, d_grey_44_end_a+1):
                                if i == int(abcd[3]):
                                    return ('grey')
                        elif int(abcd[2])==int(122):
                            for i in range(d_grey_122_start_a,d_grey_122_end_a+1):
                                if i==int(abcd[3]):
                                    return ('grey')
                        return c_key

    if inp.lower()=='sheet2' or inp=='2': 
        # Operations for sheet 2
        for row in b_sheet.iter_cols(min_col=1, min_row=4, max_col=4,max_row=100):  # iterate through all rows in specific column
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
            for cell in row:
                if cell.value != None:
                    if '/' not in inp_ip:
                        abcd = inp_ip.split('.')
                    else:
                        remove_network = inp_ip.split('/')
                        abcd = remove_network[0].split('.')
                    if inp_ip in bluezone or float(abcd[0] + '.' + abcd[1]) != ab:
                        return ('blue')
                    else:
                        if float(abcd[0] + '.' + abcd[1]) == ab:

                            c_key = get_key(int(abcd[2]), c_dict_b)
<<<<<<< HEAD

=======
                            
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                            if int(abcd[2]) == int('169'):
                                if int(abcd[3]) in range(d_green_169_start_b, d_green_169_end_b + 1):
                                    return ('green')
                            elif int(abcd[2]) == int(172):
                                for i in range(d_grey_172_1_start_b, d_grey_172_1_end_b + 1):
                                    if i == int(abcd[3]):
                                        return ('grey')
                                for i in range(d_grey_172_2_start_b, d_grey_172_2_end_b + 1):
                                    if i == int(abcd[3]):
                                        return ('grey')
                            return c_key

<<<<<<< HEAD

def parse_user_input(ss):
=======
def parse_user_input(ss): 
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
    """
    parses the source and destination ip's from the user's spreadsheet
    Input: Spreadsheet
    Output: None (Saves spreadsheet)
    """
    load_spreadsheet = openpyxl.load_workbook(ss)
    sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
<<<<<<< HEAD
    for row in sheet.iter_cols(min_col=4, min_row=2, max_col=6,max_row=100):  # iterate through all rows in specific column
        # print (type(row))
        for i in range(2, 101):
            if row == ('Sheet1.E' + str(i)):
                # print (row)
                pass
        for cell in row:
            # print (type(cell.value))
            try:
                if (cell.value) == None:
=======
    for row in sheet.iter_cols(min_col=4, min_row=2, max_col=6,
                                max_row=100):  # iterate through all rows in specific column
        #print (type(row))
        for i in range(2,101):
            if row==('Sheet1.E'+str(i)):
                #print (row)
                pass
        for cell in row:
            #print (type(cell.value))
            try:
                if (cell.value)==None:
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                    pass
                elif ',' in cell.value:
                    if cell.value.split(','):
                        for i in cell.value.split(','):
                            i = i.strip()
<<<<<<< HEAD
                            # cell_values.append(str(i))
                            if valid_ip.is_valid(i):
                                returned_zone = check_zone_ip_relationship(i, inp)
                                color_ip(returned_zone, cell)  # calling the function color_ip
                elif cell.value == None:
=======
                            #cell_values.append(str(i))
                            if is_valid(i):
                                returned_zone=check_zone_ip_relationship(i,inp)
                                color_ip(returned_zone,cell) #calling the function color_ip
                elif cell.value==None:
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                    pass
                else:
                    if cell.value.split('\n'):
                        for i in cell.value.split('\n'):
                            i = i.strip()
<<<<<<< HEAD
                            # print ('input'+i)
                            # print (inp)
                            if valid_ip.is_valid(i):
                                returned_zone = check_zone_ip_relationship(i, inp)
                                # print (returned_zone)
                                color_ip(returned_zone, cell)
            except AttributeError:
                returned_zone = check_zone_ip_relationship((cell.value), inp)
                color_ip(returned_zone, cell)
=======
                            #print ('input'+i)
                            #print (inp)
                            if is_valid(i):
                                returned_zone=check_zone_ip_relationship(i,inp)
                            #print (returned_zone)
                                color_ip(returned_zone,cell)
            except AttributeError:
                returned_zone=check_zone_ip_relationship((cell.value),inp)
                color_ip(returned_zone,cell)
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
            except ValueError:
                break
    load_spreadsheet.save(ss)

<<<<<<< HEAD

def color_ip(prevsymbol, cell):
    # colors the ip based on zone
=======
def color_ip(prevsymbol,cell): 
    #colors the ip based on zone
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
    if prevsymbol != None:
        if (prevsymbol.lower()) == "green":
            cell.fill = PatternFill(fgColor="0080FF00", fill_type="solid")
        if (prevsymbol.lower()) == "blue":
            cell.fill = PatternFill(fgColor="000080FF", fill_type="solid")
        if (prevsymbol.lower()) == "yellow":
            cell.fill = PatternFill(fgColor="00FFFF33", fill_type="solid")
        if (prevsymbol.lower()) == "red":
            cell.fill = PatternFill(fgColor="00FF3333", fill_type="solid")
        if (prevsymbol.lower()) == "grey":
            cell.fill = PatternFill(fgColor="00A0A0A0", fill_type="solid")
        if (prevsymbol.lower()) == "brown":
            cell.fill = PatternFill(fgColor="00994C00", fill_type="solid")
        if (prevsymbol.lower()) == "orange":
            cell.fill = PatternFill(fgColor="00FF9933", fill_type="solid")

<<<<<<< HEAD

=======
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
def zone(ss):
    """
    Coloring function for zone.
    """
    load_spreadsheet = openpyxl.load_workbook(ss)
    sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
    for row in sheet.iter_cols(min_col=1, min_row=1, max_col=2,
<<<<<<< HEAD
                               max_row=100):  # iterate through all rows in specific column
        for cell in row:
            prevsymbol = cell.value
            if prevsymbol != None:
                color_ip(prevsymbol, cell)
    for row in sheet.iter_cols(min_col=10, min_row=2, max_col=10,
                               max_row=100):  # iterate through all rows in specific column
=======
                                max_row=100):  # iterate through all rows in specific column
        for cell in row:
            prevsymbol = cell.value
            if prevsymbol != None:
                color_ip(prevsymbol,cell)
    for row in sheet.iter_cols(min_col=10, min_row=2, max_col=10,
                                max_row=100):  # iterate through all rows in specific column
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
        for cell in row:
            prevsymbol = cell.value
            if prevsymbol != None:
                if (prevsymbol.lower()) == "allow":
                    cell.fill = PatternFill(fgColor="0080FF00", fill_type="solid")
                if (prevsymbol.lower()) == "deny":
                    cell.fill = PatternFill(fgColor="00FF3333", fill_type="solid")
    load_spreadsheet.save(ss)

<<<<<<< HEAD

=======
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
def app_name(ss):
    """
    Coloring function based on app name.
    """
    load_spreadsheet = openpyxl.load_workbook(ss)
    sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
<<<<<<< HEAD
    app_values = []
    for row in sheet.iter_cols(min_col=9, min_row=2, max_col=9, max_row=100):
        # iterate through all rows in specific column
        for cell in row:  # add all the apps in a list app_values using different delimiters
=======
    app_values=[]
    for row in sheet.iter_cols(min_col=9, min_row=2, max_col=9, max_row=100):  
        # iterate through all rows in specific column
        for cell in row: #add all the apps in a list app_values using different delimiters
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
            try:
                if cell.value.split(','):
                    for i in cell.value.split(','):
                        app_values.append(str(i))
                elif cell.value.split(' '):
                    for i in cell.value.split(' '):
                        app_values.append(str(i))
            except AttributeError:
                app_values.append(str(cell.value))
<<<<<<< HEAD
        app_values = set(app_values)
=======
        app_values=set(app_values)
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
        app_values.remove('None')
        ft = Font(color="FF0000")
        for i in app_values:
            if i in common_apps:
                for cell in row:
                    try:
                        if cell.value.split(','):
                            for i in cell.value.split(','):
<<<<<<< HEAD
                                i.font = ft
                        elif cell.value.split(' '):
                            for i in cell.value.split(' '):
                                i.font = ft
                    except AttributeError:
                        cell.font = ft
    load_spreadsheet.save(ss)


# try to put the App:port in the same row where the corresponding app and port are found, I have just made a new row and added from top ie row 1
def port_number(ss):
    load_spreadsheet = openpyxl.load_workbook(ss)
    sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
    for i in range(7, 8):
        for k in range(50):
            ini_cell = 3
            cellValue = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
            sheet.cell(1, 12).value = 'Apps_Found (App:Port_Number)'
            if (sheet.cell(1, 12).value.lower()) == "apps_found (app:port_number)":
                sheet.cell(1, 12).fill = PatternFill(fgColor="00FFC080", fill_type="solid")
    for row in sheet.iter_cols(min_col=7, min_row=2, max_col=8,
                               max_row=100):  # iterate through all rows in specific column
        cell_values = []
        for cell in row:  # add all the ports in a list cell_values using different delimiters
            try:
                if cell.value.split(','):
                    for i in cell.value.split(','):
                        i = i.strip()
                        cell_values.append(str(i))
                elif cell.value.split('\n'):
                    for i in cell.value.split('\n'):
                        i = i.strip()
                        cell_values.append(str(i))
=======
                                i.font=ft
                        elif cell.value.split(' '):
                            for i in cell.value.split(' '):
                                i.font=ft
                    except AttributeError:
                        cell.font=ft
    load_spreadsheet.save(ss)


#try to put the App:port in the same row where the corresponding app and port are found, I have just made a new row and added from top ie row 1
def port_number(ss):
    load_spreadsheet = openpyxl.load_workbook(ss)
    sheet = load_spreadsheet['Sheet1']  # Get a sheet from the workbook
    for i in range(7,8):
        for k in range(50):
            ini_cell=3
            cellValue = str(sheet[get_column_letter(i + 1) + str(k + 1)].value)
            sheet.cell(1, 12).value = 'Apps_Found (App:Port_Number)'
            if (sheet.cell(1, 12).value.lower()) == "apps_found (app:port_number)":
                sheet.cell(1,12).fill = PatternFill(fgColor="00FFC080", fill_type="solid")
    for row in sheet.iter_cols(min_col=7, min_row=2, max_col=8,
                                    max_row=100):  # iterate through all rows in specific column
        cell_values = []
        for cell in row: #add all the ports in a list cell_values using different delimiters
            try:
                if cell.value.split(','):
                    for i in cell.value.split(','):
                        i=i.strip()
                        cell_values.append(str(i))
                elif cell.value.split('\n'):
                        for i in cell.value.split('\n'):
                            i=i.strip()
                            cell_values.append(str(i))
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                else:
                    cell_values.append(str(cell.value))
            except AttributeError:
                cell_values.append(str(cell.value))
        for i in cell_values:
            if '\n' in i:
                for j in i.split('\n'):
                    cell_values.append(str(j))
<<<<<<< HEAD
        cell_values = set(cell_values)
        cell_values.remove('None')
        apps_found = {}
        for prevsymbol in cell_values:  # looping each port in cell_values list
            if prevsymbol != None:
                prevsymbol = str(prevsymbol)
                for key, value in common_apps.items():
                    if prevsymbol == value:
                        apps_found[key] = prevsymbol
=======
        cell_values=set(cell_values)
        cell_values.remove('None')
        apps_found={}
        for prevsymbol in cell_values: #looping each port in cell_values list
            if prevsymbol!=None:
                prevsymbol=str(prevsymbol)
                for key,value in common_apps.items():
                    if prevsymbol == value:
                        apps_found[key]=prevsymbol
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
                    else:
                        list_of_keys = [key
                                        for key, list_of_values in common_apps_list.items()
                                        if prevsymbol in list_of_values]
                        if list_of_keys:
                            for i in set(list_of_keys):
                                if i not in apps_found:
                                    apps_found[i] = prevsymbol
<<<<<<< HEAD
                if prevsymbol not in (
                common_apps):  # idea is to create a pool of service objects, so that we can highlight them the next time they are seen
                    common_apps['tcp/udp-' + str(prevsymbol)] = str(prevsymbol)
            else:
                pass
        ini_cell = 3
        already_added = []
        for key, value in apps_found.items():
=======
                if prevsymbol not in (common_apps): # idea is to create a pool of service objects, so that we can highlight them the next time they are seen
                    common_apps['tcp/udp-'+str(prevsymbol)]=str(prevsymbol)
            else:
                pass
        ini_cell=3
        already_added=[]
        for key,value in apps_found.items():
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
            if value in apps_found.values():
                if value not in already_added:
                    sheet.cell((ini_cell), 12).value = key + ':' + value
                    sheet.cell((ini_cell), 12).fill = PatternFill(fgColor="00FFE0C0", fill_type="solid")
                    already_added.append(value)
<<<<<<< HEAD
                    ini_cell += 1
=======
                    ini_cell+=1
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
    load_spreadsheet.save(ss)


def main():
<<<<<<< HEAD
    # a.b.c.d-> ip format
    # Sheet1 subnets

    # Use valid_ip.py in repo as module and import it in this script. Use OOPs concepts to eliminate below function

    # I am using the below two lists here which are kinda hard coded, I want to identify applications from https://applipedia.paloaltonetworks.com/ based on the input port number
=======
    #a.b.c.d-> ip format
    #Sheet1 subnets
    
    # Use valid_ip.py in repo as module and import it in this script. Use OOPs concepts to eliminate below function
 
    #I am using the below two lists here which are kinda hard coded, I want to identify applications from https://applipedia.paloaltonetworks.com/ based on the input port number
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8

    global inp

    inp = input('Which Device? Type 1 for sheet1 or 2 for sheet2 :\n')
    ss_inp = input('Enter the path to excel sheet\n')
    parse_user_input(ss_inp)

    zone(ss_inp)
    port_number(ss_inp)

<<<<<<< HEAD
    # app_name(ss_inp)
if __name__ == "__main__":
    main()
=======
    #app_name(ss_inp)
    

if __name__== "__main__":
    main()
>>>>>>> 289eec883f2577a7ee2835b40e96128e45e012e8
