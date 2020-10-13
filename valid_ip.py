import openpyxl
import ipaddress  # included in Python 3
from pprint import pprint   #to print in a human readable format

# i=input('Enter the name of file:\nMake sure the file is in the same folder as this python file, \
#otherwise input the file address as "C:/Users/Lenovo/Downloads/workbook.xlsx"\n')
i = ('Hacktoberfest_Inputt.xlsx')


# function to validate a IP address.
def is_valid(ip):
    try:
        ipaddress.IPv4Network(ip)
        return True
    except ValueError:
        return False

# for processing the string before validating.
def process(s):
    
    # removing delimeters and commas.
    for idx, k in enumerate(s):
        if k=='\n' or k==',':
            s = s[:idx] + " " + s[idx + 1:]
    s = s.split()
    
    # for processing the networks like "12.34.54.67/34"
    for index, i in enumerate(s):
        
        # for counting the number of full stops (must be 3) before
        # removing the slash "/" and post string.
        count=0
        for idx, k in enumerate(i):
            if k==".":
                count+=1
            if count==3 and k=="/":
                count=-1
                break
        if count==-1:
            s[index] = s[index][:idx]
    
    # removes everything except full stops and numbers.
    for index, i in enumerate(s):
        for idx, k in enumerate(i):
            if not (k == '.' or k.isnumeric()):
                s[index] = s[index][:idx] + " " + s[index][idx + 1:]

        s[index] = s[index].replace(" ","")
    
    return(s)

# for processing the IP addresses just before adding them 
# to the dictionary.
def add_process(s):
    
    # removing delimeters and commas.
    for idx, k in enumerate(s):
        if k=='\n' or k==',':
            s = s[:idx] + " " + s[idx + 1:]
    s = s.split()
    
    return(', '.join(s))

# Driving Code
if __name__ == "__main__":
    
    wb = openpyxl.load_workbook(i)

    ws = wb['Sheet1']

    col=[3,5]   # INPUT. A list containing the indices of the
                # columns, which needs to be checked for the
                # IP adresses.

    lis=[]      # for storing the valid cell values.
                # It will be a list of dictionaries with keys as
                # the respective column headings.

    # 2 is so as to ignore the headings row.
    for i in range(2, ws.max_row+1):
        
        # found variable checks if any of the column consists of
        # non IP address format values.
        found=False
        
        for j in col:

            if ws[i][j].value:

                s = process(ws[i][j].value)
                
                for k in s:
                    if not is_valid(k):
                        found = True
                        break
                if found: break
                    
            else:
                found = True
 
        if not found:
            dic = {}
            for j in range(0,len(col)-1,2):
                # dic[add_process(ws[1][j].value)] = [add_process(ws[i][j].value)]
                dic[add_process(ws[i][col[j]].value)] = [add_process(ws[i][col[j+1]].value)]
            lis.append(dic)
        # print("next loop")
    print(" ")
    
    pprint(lis)
