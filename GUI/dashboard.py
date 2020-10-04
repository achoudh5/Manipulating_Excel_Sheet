from tkinter import *
#import <api>
from PIL import ImageTk, Image

from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook

from tkinter.font import Font
from tkinter.ttk import Style

# setting the path to parent directory so as to import zip
import sys
sys.path.insert(0,'..')

from zip import main

root = Tk()

root.title("Check Validity")

root.geometry("1250x700")


## Adding ScrollBar

main_frame = Frame(root)
main_frame.pack(fill=BOTH, expand=1)

my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

style = ttk.Style()
# scrlbr.configure('scrlbr1', troughcolor='black')

# style.element_create("My.Vertical.Scrollbar.trough", "from", "default")
style.element_create("My.Vertical.Scrollbar.trough", "from", "default")

# Redefine the vertical scrollbar layout to use the custom trough.
style.layout("My.Vertical.TScrollbar",
    [('My.Vertical.Scrollbar.trough', {'children':
        [('Vertical.Scrollbar.uparrow', {'side': 'top', 'sticky': ''}),
         ('Vertical.Scrollbar.downarrow', {'side': 'bottom', 'sticky': ''}),
         ('Vertical.Scrollbar.thumb', {'unit': '1', 'children':
             [('Vertical.Scrollbar.grip', {'sticky': ''})],
        'sticky': 'nswe'})],
    'sticky': 'ns'})])

# Copy original style configuration and add our new custom configuration option.
style.configure("My.Vertical.TScrollbar", troughcolor="black")

my_scrollbar_ver = ttk.Scrollbar(main_frame, orient=VERTICAL, command = my_canvas.yview, style="My.Vertical.TScrollbar")
my_scrollbar_ver.pack(side=RIGHT, fill=Y)

my_canvas.configure(yscrollcommand=my_scrollbar_ver.set)
my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion = my_canvas.bbox("all")))

# this second_frame is the new 'root' now.
second_frame = Frame(my_canvas)

my_canvas.create_window((0,0), window=second_frame, anchor="nw")


def fileDownload(f):
	f.save('output.xlsx')

def fileDialog():
	filename = filedialog.askopenfilename(initialdir = "/", title = "Select an Excel File", filetype = (("xlsx", "*.xlsx"), ("All Files", "*.*")))
	label_3.config(text=filename)

	try:

		if e.get().isnumeric():
			main(filename, int(e.get()))
		else:
			raise ValueError

		wb = load_workbook(filename)
		sh = "Sheet"
		sh = sh + str(e.get())
		sheet = wb[sh]

		button1 = ttk.Button(second_frame, text="Download", command=fileDownload(wb))
		button1.grid(sticky='W', padx=65, pady=10)

		treeview["columns"] = [sheet[1][i].value for i in range(sheet.max_column)]
		treeview["show"] = "headings"

		for i in range(sheet.max_column):
			treeview.heading(sheet[1][i].value, text=sheet[1][i].value)

		index = 0
		for row in range(2, sheet.max_row+1):
			ls=[]

			for col in range(sheet.max_column):
				ls.append(sheet[row][col].value)

			ls = tuple(ls)

			treeview.insert("", "end", values = ls)

			index += 1

	except:

		label_3.config(text="Some Error occured. Confirm the sheet number in the particular file or check the file format (.xlsx).")

########## original dashbboard code STARTS


# canvas1 = Canvas(root, width = 400, height = 300)
# canvas1.grid()
# The Canvas is your display where you can place items, such as entry boxes, buttons, charts and more.
# You can control the dimensions of your Canvas by changing the width and height values:


# my_img = ImageTk.PhotoImage(Image.open("../images/Screen Shot 2020-10-03 at 1.19.35 AM.png"))

image = Image.open('../images/Screen Shot 2020-10-03 at 1.19.35 AM.png')
# The (450, 350) is (height, width)
image = image.resize((490, 390), Image. ANTIALIAS)
my_img = ImageTk.PhotoImage(image)

my_label = Label(second_frame,image=my_img)
my_label.grid(column=2, row=1)

def myClick():
    global counter
    counter+=1
    #s= <api call happening, this is already in place, check import>
    s="abcd@gmail.com"  # for testing purposes
    mylabel = Label(second_frame, text=str(counter)+' '+'You Entered: '\
                               +input_box.get()+'\nRequestor: '+' '+str(s)+'\n')
    mylabel.grid(column=3, row=6)
    input_box.delete(0, END)

#counts how many time you used this tool to check validity
counter=0

#An input_box can be used to get the user’s input
input_box = Entry(second_frame, width=40, bg='#30CFBB', fg='black', selectforeground='red',selectbackground='white',borderwidth=5)
# canvas1.create_window(250,140,window=input_box)


label2 = Label(second_frame, text='*The database is case sensitive!*')
label2.grid(column=3, row=3, padx=50)
label2.config(font=('helvetica', 16))
# canvas1.create_window(250, 180, window=label2)
input_box.grid(column=3, row=4, padx=50)
#input_box.insert() #placeholder text in the box

#validity_Button = Button(root, text="Checking validity of the Requestor!", padx=70, pady=70, command=lambda: myClick(bluegroup_api.Bluegroup(input_box.get())))
validity_Button = Button(second_frame,text="Checking validity!",command= myClick, bg='brown', fg='white', font=('helvetica', 9, 'bold'), cursor="hand2")
# canvas1.create_window(280, 180, window=validity_Button)
validity_Button.grid(column=3, row=5, padx=50)

#Exit Button
#button_quit = Button(root, text="Exit Program", command=root.quit)
#button_quit.grid()


######### original dashbboard code ENDS

label = ttk.Label(second_frame, text="INPUT:", font=('Arial', 15, 'bold'))
label.grid(column=0, row=2, pady=(10,0), sticky='W', padx=30)

label_1 = ttk.Label(second_frame, text="Type the sheet number of the Excel file :", font=('Arial', 12))
label_1.grid(column=0, row=3, padx=35, pady=(15,5), sticky='W')

label_4 = ttk.Label(second_frame, text="Type 1 for Sheet 1 or Type 2 for Sheet 2", font=('Arial', 10))
label_4.grid(column=0, row=4, padx=35, pady=(0,10), sticky='W')

e = Entry(second_frame, width=40, border=3)
e.grid(column=0, row=5, pady=(0,5), sticky='W', padx=35)

label_2 = ttk.Label(second_frame, text="Open a File :", font=('Arial', 12))
label_2.grid(column=0, row=6, padx=35, pady=15, sticky='W')

button = Button(second_frame, text="Browse File", command=fileDialog, bg='brown', fg='white', font=('helvetica', 9, 'bold'), cursor='hand2')
button.grid(column=0, row=7, sticky='W', padx=35)

label_3 = ttk.Label(second_frame, text="", font=('Arial', 10))
# label_3.grid(column=1, row=5)
label_3.grid(sticky='W', padx=35)

label_5 = ttk.Label(second_frame, text="OUTPUT:", font=('Arial', 15, 'bold'))
label_5.grid(pady=(30,10), sticky='W', padx=20)

font=Font(family='Arial', size=20)
fontheight=font.metrics()['linespace']
style=Style()
style.configure('Calendar.Treeview', rowheight=fontheight, font=('Arial', 11)) 

treeview = ttk.Treeview(second_frame, style='Calendar.Treeview')

treeview.grid(padx=20, pady=0)

mainloop()
