from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import load_workbook
from zip import main

from tkinter.font import Font
from tkinter.ttk import Style

root = Tk()
root.title("Parsing Excel Sheet")
root.geometry("1250x700")

def fileDownload(f):
	f.save('output.xlsx')

def fileDialog():
	filename = filedialog.askopenfilename(initialdir = "C:/Users/Lenovo/Desktop", title = "Select an Excel File", filetype = (("xlsx", "*.xlsx"), ("All Files", "*.*")))
	label_3.config(text=filename)

	try:

		if int(e.get()).isnumeric():
			main(filename, e.get())
		else:
			raise ValueError

		wb = load_workbook(filename)
		sh = "Sheet"
		sh = sh + str(e.get())
		sheet = wb[sh]

		button1 = ttk.Button(root, text="Download", command=fileDownload(wb))
		button1.grid(sticky='W', padx=20)

		treeview["columns"] = [sheet[1][i].value for i in range(sheet.max_column)]
		treeview["show"] = "headings"

		for i in range(sheet.max_column):
			treeview.heading(sheet[1][i].value, text=sheet[1][i].value)

		index = 0
		for row in range(2, sheet.max_row+1):
			ls=[]

			for col in range(sheet.max_column):
				ls.append(sheet[row][col].value)

			ls = tuple(ls)

			treeview.insert("", "end", values = ls)

			index += 1

	except:

		label_3.config(text="Some Error occured. Confirm the sheet number in the particular file or check the file format (.xlsx).")


label = ttk.Label(root, text="INPUT:", font=('Arial', 15, 'bold'))
label.grid(pady=(10,0), sticky='W', padx=5)

label_1 = ttk.Label(root, text="Type the sheet number of the Excel file :", font=('Arial', 11))
label_1.grid(padx=20, pady=(15,5), sticky='W')

label_4 = ttk.Label(root, text="Type 1 for Sheet 1 or Type 2 for Sheet 2", font=('Arial', 9))
label_4.grid(padx=20, pady=(0,10), sticky='W')

e = Entry(root, width=40, border=3)
e.grid(pady=(0,5), sticky='W', padx=20)

label_2 = ttk.Label(root, text="Open a File :", font=('Arial', 11))
label_2.grid(padx=20, pady=15, sticky='W')


button = ttk.Button(root, text="Browse File", command=fileDialog)
button.grid(column=0, row=4, sticky='W', padx=20)

label_3 = ttk.Label(root, text="", font=('Arial', 10))
label_3.grid(column=1, row=5)

label_5 = ttk.Label(root, text="OUTPUT:", font=('Arial', 15, 'bold'))
label_5.grid(pady=(30,10), sticky='W', padx=5)

font=Font(family='Arial', size=20)
fontheight=font.metrics()['linespace']
style=Style()
style.configure('Calendar.Treeview', rowheight=fontheight) 

treeview = ttk.Treeview(root, style='Calendar.Treeview')

treeview.grid(padx=20, pady=0)

mainloop()
